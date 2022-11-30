from typing import Union

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

mappings = {
    'RCGIS DROPS': 'DROPS',
    'DIALOGTJENESTE-API': 'DIALOGTJENESTE',
    'DIALOGTJENESTE-UT':  'DIALOGTJENESTE',
    'EGET LØP': 'PAGERO',
    'IFS': 'IFS CLOUD',
    'IFSCLOUD': 'IFS CLOUD',
    'MCPS-INNBETALINGSFILER': 'MCPS',
    'MCPS-REMITTERINGSFILER': 'MCPS',
    'NETBAS-MÅLEPUNKT-UT': 'NETBAS',
    'NETBAS-SAMLESKINNENAVN (SØR)': 'NETBAS',
    'NETBAS-TILKNYTNINGSPUNKT-UT': 'NETBAS',
    'PAGERO ?': 'PAGERO',
    'STATNETT-MARGINALTAPSNAVN (NORD)': 'STATNETT',
    'STATNETT-MARGINALTAPSSATSER (NORD)': 'STATNETT',
#    'UN': 'ARCGIS',
    '????': '????????'
}

#versions = ['DV4/5', 'DV6', 'DV7', 'DVStange']
#versions = ['DV7']
versions = ['DV4/5']

# System in column A
system_column = 1
#Direction in column C
direction_column = 3
# Dependent om API in column D
api_column = 4
# Version in column AF
version_column = 32

def get_apis(api_string):
    funcs = []
    a_name = ''
    split_position = 0
    the_name = ''
    apis = []
    if type(api_string) == str:
        if ',' in api_string or '\n' in api_string:
            api_string = api_string.replace('\n', ', ')
            funcs = list(map(str.strip, api_string.split(',')))
#            funcs = list(map(str.strip, api_string.split()))
        else:
            #                print(api)
            funcs = [api_string]
        for a_name in funcs:
            a_name = a_name.upper()
            if '-' in a_name:
                split_position = a_name.index('-')
                the_name = a_name[:split_position]
                if the_name in mappings:
                    the_name = mappings[the_name]
            #                    if name not in systems:
            #                       name = api_name
            else:
                the_name = a_name
                if the_name in mappings:
                    the_name = mappings[the_name]
            if the_name != '':
                apis.append(the_name)
        if len(apis) >= 1:
            return apis
        else:
            return None
    return None


if __name__ == '__main__':

    EXCEL_WORKBOOK_NAME = '/users/djr/Downloads/DV4_Status_Integrasjoner (16).xlsx'
    EXCEL_SHEET_NAME = 'Integrasjon-Dataflyt status'

    workbook = load_workbook(EXCEL_WORKBOOK_NAME)
    worksheet = workbook[EXCEL_SHEET_NAME]

    systems = set()
    startrow = 2
    endrow = worksheet.max_row + 1
    print(endrow)
    name = ''
    version = ''
    for i in range(startrow, endrow):
        version = worksheet.cell(row=i, column=32).value
        if not version or version == '' or version in versions:
            name = worksheet.cell(row=i, column=1).value
            if name:
                systems.add(name.upper())

    functions = []
    api = ''
    for i in range(startrow, endrow):
        functions = get_apis(worksheet.cell(row=i, column=4).value)
        if functions:
            for api in functions:
                systems.add(api)

    sorted_systems = list(sorted(systems))
#    print(sorted_systems)

#    sorted_apis = list(sorted(dependent_on_apis))
#    print(sorted(sorted_apis))

    integrations_overview_sheet = workbook.create_sheet(title='Integrasjoner - oversikt')
    integrations_table_sheet = workbook.create_sheet(title='Integrasjoner - tabell')

    the_row = 1
    the_col = 2
    for name in sorted_systems:
        integrations_overview_sheet.cell(row=the_row, column=the_col).value = name
        the_col += 1

    the_row = 2
    the_col = 1
    for name in sorted_systems:
        integrations_overview_sheet.cell(row=the_row, column=the_col).value = name
        the_row += 1

    center_alignment = Alignment(horizontal='center')
    for i in range(startrow, endrow):
        version = worksheet.cell(row=i, column=32).value
        if not version or version == '' or version in versions:
            name = worksheet.cell(row=i, column=1).value
            direction = worksheet.cell(row=i, column=3).value
            if name:
                name = name.upper()
                functions = get_apis(worksheet.cell(row=i, column=4).value)
                if functions:
                    for api in functions:
                        if direction != 'Inn':
                            integrations_overview_sheet.cell(row=sorted_systems.index(name) + 2, column=sorted_systems.index(api) + 2).value = 'X'
                            integrations_overview_sheet.cell(row=sorted_systems.index(name) + 2, column=sorted_systems.index(api) + 2).alignment = center_alignment
                        if direction != 'Ut':
                            integrations_overview_sheet.cell(row=sorted_systems.index(api) + 2, column=sorted_systems.index(name) + 2).value = 'X'
                            integrations_overview_sheet.cell(row=sorted_systems.index(api) + 2, column=sorted_systems.index(name) + 2).alignment = center_alignment


    grey_fill = PatternFill(fill_type='solid', start_color='DDDDDD', end_color='DDDDDD')
    for i in range(2, len(sorted_systems)+2):
        integrations_overview_sheet.cell(row=i, column=i).fill = grey_fill

    grey_fill = PatternFill(fill_type='solid', start_color='EEEEEE', end_color='EEEEEE')
    integrations_table_sheet.cell(row=1, column=1).value = "Fra"
    integrations_table_sheet.cell(row=1, column=1).fill = grey_fill
    integrations_table_sheet.cell(row=1, column=2).value = "Til"
    integrations_table_sheet.cell(row=1, column=2).fill = grey_fill
    integrations_table_sheet.cell(row=1, column=3).value = "Forbindelse"
    integrations_table_sheet.cell(row=1, column=3).fill = grey_fill
    integrations_table_sheet.cell(row=1, column=4).value = "Krav"
    integrations_table_sheet.cell(row=1, column=4).fill = grey_fill
    integrations_table_sheet.cell(row=1, column=5).value = "Kommentar"
    integrations_table_sheet.cell(row=1, column=5).fill = grey_fill
    the_row = 2

    for i in range(2, len(sorted_systems)+2):
        for j in range(2, len(sorted_systems)+2):
            if integrations_overview_sheet.cell(row=i, column=j).value == "X":
                integrations_table_sheet.cell(row=the_row, column=1).value = sorted_systems[i-2]
                integrations_table_sheet.cell(row=the_row, column=2).value = sorted_systems[j-2]
                the_row += 1


    workbook.save(filename='/users/djr/Downloads/foo.xlsx')



